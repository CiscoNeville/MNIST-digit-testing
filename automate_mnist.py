#!/usr/bin/env python3
"""
Automated script to run MNIST training with different configurations
and collect accuracy results across multiple epochs.
"""

import subprocess
import re
import pandas as pd
import time
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

def run_mnist_training(epochs, train_classes, train_samples=50000, test_samples=10000, 
                      batch_size=128, train_seed=1, test_seed=1):
    """
    Run the MNIST training script with specified parameters.
    
    Returns:
        float: Overall accuracy percentage, or None if failed to parse
    """
    cmd = [
        "./mnist_digits_10.py",
        "--train_samples", str(train_samples),
        "--test_samples", str(test_samples),
        "--epochs", str(epochs),
        "--batch_size", str(batch_size),
        "--train_seed", str(train_seed),
        "--test_seed", str(test_seed),
        "--train_classes", train_classes
    ]
    
    try:
        print(f"Running: {' '.join(cmd)}")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=3600)
        
        if result.returncode != 0:
            print(f"Error running command: {result.stderr}")
            return None
        
        # Parse the output to extract overall accuracy
        output = result.stdout
        accuracy_match = re.search(r"Overall accuracy on \d+ test images: ([\d.]+)%", output)
        
        if accuracy_match:
            accuracy = float(accuracy_match.group(1))
            print(f"  → Accuracy: {accuracy}%")
            return accuracy
        else:
            print("  → Failed to parse accuracy from output")
            return None
            
    except subprocess.TimeoutExpired:
        print(f"  → Timeout after 1 hour")
        return None
    except Exception as e:
        print(f"  → Error: {e}")
        return None

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Automated MNIST training with different configurations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  ./automate_mnist.py --epochs 50
  ./automate_mnist.py --epochs 25 --batch_size 64
  ./automate_mnist.py --epochs 100 --train_samples 30000
        """
    )
    
    parser.add_argument(
        "--epochs", 
        type=int, 
        default=100,
        help="Number of epochs to train (default: 100)"
    )
    
    parser.add_argument(
        "--batch_size",
        type=int,
        default=128,
        help="Batch size for training (default: 128)"
    )
    
    parser.add_argument(
        "--train_samples",
        type=int,
        default=50000,
        help="Number of training samples (default: 50000)"
    )
    
    parser.add_argument(
        "--test_samples",
        type=int,
        default=10000,
        help="Number of test samples (default: 10000)"
    )
    
    parser.add_argument(
        "--train_seed",
        type=int,
        default=1,
        help="Random seed for training (default: 1)"
    )
    
    parser.add_argument(
        "--test_seed",
        type=int,
        default=1,
        help="Random seed for testing (default: 1)"
    )
    
    return parser.parse_args()

def main():
    # Parse command line arguments
    args = parse_arguments()
    
    # Define the training class configurations
    configurations = [
        ("d,w,nan", "Config 1: d,w,nan"),
        ("d,w", "Config 2: d,w"),
        ("d,nan", "Config 3: d,nan"),
        ("d", "Config 4: d")
    ]
    
    # Initialize results dictionary and timing tracking
    results = {}
    
    # Track progress and timing
    total_runs = len(configurations) * args.epochs
    current_run = 0
    start_time = time.time()
    epoch_times = []  # Track individual epoch times
    
    print(f"Starting automated MNIST training...")
    print(f"Total runs planned: {total_runs}")
    print(f"Epochs per configuration: {args.epochs}")
    print(f"Batch size: {args.batch_size}")
    print(f"Train samples: {args.train_samples}")
    print(f"Test samples: {args.test_samples}")
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # Run each configuration
    for train_classes, config_name in configurations:
        print(f"\n{config_name}")
        print("-" * 40)
        
        config_results = []
        
        # Run epochs 1 to specified number
        for epoch in range(1, args.epochs + 1):
            current_run += 1
            epoch_start_time = time.time()
            
            print(f"[{current_run}/{total_runs}] Epoch {epoch} with {train_classes}")
            
            accuracy = run_mnist_training(
                epochs=epoch,
                train_classes=train_classes,
                train_samples=args.train_samples,
                test_samples=args.test_samples,
                batch_size=args.batch_size,
                train_seed=args.train_seed,
                test_seed=args.test_seed
            )
            
            epoch_end_time = time.time()
            epoch_duration = epoch_end_time - epoch_start_time
            epoch_times.append(epoch_duration)
            config_results.append(accuracy)
            
            # Progress update every 10 epochs (or every 5 if less than 20 total epochs)
            progress_interval = 5 if args.epochs < 20 else 10
            if epoch % progress_interval == 0:
                completed = sum(1 for x in config_results if x is not None)
                failed = epoch - completed
                
                # Calculate time per epoch based on cumulative nature
                # Each epoch N takes roughly N times as long as epoch 1
                if len(epoch_times) >= 2:
                    # Estimate time per "training epoch" from recent data
                    recent_times = epoch_times[-min(progress_interval, len(epoch_times)):]
                    recent_epochs = list(range(max(1, epoch - len(recent_times) + 1), epoch + 1))
                    
                    # Calculate average time per cumulative epoch
                    total_recent_time = sum(recent_times)
                    total_recent_epochs_trained = sum(recent_epochs)  # Total epochs actually trained
                    time_per_training_epoch = total_recent_time / total_recent_epochs_trained if total_recent_epochs_trained > 0 else 0
                    
                    # Estimate remaining time for this configuration
                    remaining_epochs_this_config = list(range(epoch + 1, args.epochs + 1))
                    remaining_training_epochs_this_config = sum(remaining_epochs_this_config)
                    
        
        # Store results for this configuration
        results[config_name] = config_results
        
        # Configuration summary
        completed_in_config = sum(1 for x in config_results if x is not None)
        print(f"\n{config_name} completed: {completed_in_config}/{args.epochs} successful")
        if completed_in_config > 0:
            config_accuracies = [x for x in config_results if x is not None]
            print(f"Best accuracy: {max(config_accuracies):.2f}%")
        print("-" * 50)
    
    # Create final results Excel file
    create_final_excel(results, args)
    
    # Final summary
    total_duration = time.time() - start_time
    print(f"\nCompleted at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total time: {total_duration/3600:.1f} hours")
    print("All training runs completed!")

def create_final_excel(results, args):
    """Create and save the final results Excel file with betterment table."""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Accuracy Results'
    
    # Row 1: Parameters
    params_text = f"--train_samples {args.train_samples} --test_samples {args.test_samples} --epochs {args.epochs} --batch_size {args.batch_size} --train_seed {args.train_seed} --test_seed {args.test_seed}"
    ws['A1'] = params_text
    
    # Row 2-3: Blank rows
    
    # Row 4: Headers
    headers = ['Epoch', 'Config 1: d,w,nan', 'Config 2: d,w', 'Config 3: d,nan', 'Config 4: d']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Data starting from row 5
    for epoch in range(1, args.epochs + 1):
        ws.cell(row=4 + epoch, column=1).value = epoch
        
        col = 2  # Start from column B (after Epoch column)
        for config_name in results.keys():
            config_results = results[config_name]
            if epoch <= len(config_results) and config_results[epoch-1] is not None:
                ws.cell(row=4 + epoch, column=col).value = config_results[epoch-1]
            col += 1
    
    # Betterment table
    # G3: Merged header
    ws.merge_cells('G3:I3')
    betterment_header = ws['G3']
    betterment_header.value = 'Betterment'
    betterment_header.font = Font(bold=True)
    betterment_header.fill = PatternFill(start_color='808000', end_color='808000', fill_type='solid')  # Olive green
    
    # G4-I4: Column headers
    betterment_headers = ['w,nan', 'w', 'nan']
    for col, header in enumerate(betterment_headers):
        cell = ws.cell(row=4, column=7 + col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Betterment formulas starting from row 5
    betterment_cols = ['G', 'H', 'I']  # w,nan, w, nan
    reference_cols = ['B', 'C', 'D']   # d,w,nan vs d,w vs d,nan
    
    for epoch in range(1, args.epochs + 1):
        row = 4 + epoch
        for col_idx, (bet_col, ref_col) in enumerate(zip(betterment_cols, reference_cols)):
            formula = f"={ref_col}{row}-$E{row}"  # Compare to Config 4: d (column E)
            ws[f"{bet_col}{row}"] = formula
    
    # Note: Conditional formatting removed due to openpyxl API complexities
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save Excel file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f"mnist_accuracy_results_{args.epochs}epochs_{timestamp}.xlsx"
    wb.save(excel_filename)
    
    print(f"\nResults saved to: {excel_filename}")
    
    # Print summary statistics
    print(f"\nSummary Statistics:")
    print("=" * 50)
    
    for config_name in results.keys():
        config_results = [x for x in results[config_name] if x is not None]
        if config_results:
            print(f"{config_name}:")
            print(f"  Completed runs: {len(config_results)}/{args.epochs}")
            print(f"  Best accuracy: {max(config_results):.2f}%")
            final_epoch_accuracy = config_results[-1] if len(config_results) == args.epochs else 'N/A'
            print(f"  Final accuracy (epoch {args.epochs}): {final_epoch_accuracy}%")
            print(f"  Average accuracy: {sum(config_results)/len(config_results):.2f}%")
            print()

if __name__ == "__main__":
    # Check if script exists
    import os
    if not os.path.exists("./mnist_digits_10.py"):
        print("Error: mnist_digits_10.py not found in current directory")
        print("Make sure you're running this script from the same directory as mnist_digits_10.py")
        sys.exit(1)
    
    # Check if required packages are available
    try:
        import pandas as pd
        import openpyxl
    except ImportError:
        print("Error: pandas and openpyxl are required. Install with: pip install pandas openpyxl")
        sys.exit(1)
    
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nScript interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)